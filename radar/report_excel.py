#!/usr/bin/env python3
"""Professional Excel report from enriched listings + new-ids set."""
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule, CellIsRule

A = "Arial"
HF = PatternFill("solid", fgColor="1F4E79"); HFONT = Font(name=A, bold=True, color="FFFFFF", size=10)
CELL = Font(name=A, size=10); LINK = Font(name=A, size=10, color="0563C1", underline="single")
TH = Side(style="thin", color="D0D0D0"); BD = Border(left=TH, right=TH, top=TH, bottom=TH)
TOP = Alignment(vertical="top", wrap_text=True); CT = Alignment(vertical="top", horizontal="center")
BAND = PatternFill("solid", fgColor="F2F6FC"); BACK = PatternFill("solid", fgColor="FCE4D6")
NEWFILL = PatternFill("solid", fgColor="E2EFDA")

COLS = [("new", "🆕", 5), ("name", "Ad (Model)", 34), ("brand", "Brend", 11), ("band", "Qiymət aralığı", 12),
        ("price", "Qiymət", 10), ("spec_score", "Spec balı", 9), ("value_score", "Dəyər balı", 9),
        ("params", "Parametrlər", 40), ("cpu", "CPU", 16), ("ram", "RAM", 6), ("storage", "Yaddaş", 12),
        ("screen", "Ekran", 7), ("gpu", "GPU", 13), ("usage", "İstifadə", 15), ("is_new", "Yeni?", 6),
        ("seller_type", "Satıcı tipi", 11), ("seller", "Satıcı", 20), ("phones", "Telefon", 24),
        ("link", "Elan linki", 40), ("region", "Region", 10), ("updated_at", "Yenilənmə", 12)]
CKK = {"new", "price", "spec_score", "value_score", "ram", "screen"}


def _backlink(ws):
    c = ws.cell(row=1, column=1, value="⬅  Xülasəyə qayıt")
    c.hyperlink = Hyperlink(ref="A1", location="'Xülasə'!A1", display="Xülasəyə")
    c.font = Font(name=A, bold=True, size=11, color="C00000", underline="single")
    for cc in range(1, 5): ws.cell(row=1, column=cc).fill = BACK
    ws.row_dimensions[1].height = 20


def _table(ws, rows, new_ids, tab=None):
    _backlink(ws)
    if tab: ws.sheet_properties.tabColor = tab
    for i, (k, lb, w) in enumerate(COLS, 1):
        ws.cell(row=2, column=i, value=lb); ws.column_dimensions[get_column_letter(i)].width = w
        ws.cell(row=2, column=i).fill = HF; ws.cell(row=2, column=i).font = HFONT
        ws.cell(row=2, column=i).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}{len(rows)+2}"
    rr = 3
    for ridx, r in enumerate(rows):
        isnew = r["ad_id"] in new_ids
        for i, (k, lb, w) in enumerate(COLS, 1):
            if k == "new":
                v = "🆕" if isnew else ""
            else:
                v = r.get(k)
            c = ws.cell(row=rr, column=i, value=v); c.font = CELL; c.border = BD
            c.alignment = CT if k in CKK else TOP
            if isnew: c.fill = NEWFILL
            elif ridx % 2: c.fill = BAND
            if k == "price": c.number_format = "#,##0"
            if k in ("spec_score", "value_score"): c.number_format = "0.0"
            if k == "link" and v: c.hyperlink = v; c.font = LINK
            if k == "seller_type": c.font = Font(name=A, size=10, bold=True, color="2E7D32" if v == "Mağaza" else "B26A00")
            if k == "usage": c.font = Font(name=A, size=10, bold=True, color="6A1B9A" if v == "Gaming" else "1565C0")
        ws.row_dimensions[rr].height = 30; rr += 1
    last = len(rows) + 2
    if last >= 3:
        keys = [k for k, _, _ in COLS]
        gl = get_column_letter(keys.index("value_score") + 1)
        ws.conditional_formatting.add(f"{gl}3:{gl}{last}", ColorScaleRule(
            start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50,
            mid_color="FFEB84", end_type="max", end_color="63BE7B"))
        el = get_column_letter(keys.index("price") + 1)
        ws.conditional_formatting.add(f"{el}3:{el}{last}", DataBarRule(start_type="min", end_type="max", color="F4B183"))
        sc = get_column_letter(keys.index("seller_type") + 1)
        maxL = get_column_letter(len(COLS))
        ws.conditional_formatting.add(f"A3:{maxL}{last}", FormulaRule(formula=[f'${sc}3="Şəxsi"'], fill=BACK))


def build_excel(listings, new_ids, run_ts, path):
    wb = Workbook()
    ws = wb.active; ws.title = "Xülasə"; ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F4E79"

    L = sorted(listings, key=lambda r: -(r.get("value_score") or 0))
    news = sorted([r for r in L if r["ad_id"] in new_ids], key=lambda r: -(r.get("value_score") or 0))
    gaming = [r for r in L if r.get("usage") == "Gaming"]
    ofis = [r for r in L if r.get("usage") != "Gaming"]

    # data sheets
    _table(wb.create_sheet("🆕 Yeni"), news, new_ids, tab="2E7D32")
    _table(wb.create_sheet("Bütün Elanlar"), L, new_ids, tab="44546A")
    _table(wb.create_sheet("Gaming"), gaming, new_ids, tab="6A1B9A")
    _table(wb.create_sheet("Ofis"), ofis, new_ids, tab="1565C0")

    # hidden chart data
    hs = wb.create_sheet("_data"); hs.sheet_state = "hidden"
    band_c = Counter(r["band"] for r in L if r.get("band"))
    band_order = sorted(band_c, key=lambda b: int(str(b).split("–")[0]) if str(b).split("–")[0].isdigit() else 0)
    brand_c = Counter(r["brand"] for r in L if r.get("brand"))
    hs["A1"] = "Band"; hs["B1"] = "Say"
    for i, b in enumerate(band_order, 2): hs.cell(row=i, column=1, value=b); hs.cell(row=i, column=2, value=band_c[b])
    hs["D1"] = "Brend"; hs["E1"] = "Say"
    for i, (b, c) in enumerate(brand_c.most_common(10), 2): hs.cell(row=i, column=4, value=b); hs.cell(row=i, column=5, value=c)
    hs["G1"] = "İstifadə"; hs["H1"] = "Say"
    hs["G2"] = "Gaming"; hs["H2"] = len(gaming); hs["G3"] = "Ofis"; hs["H3"] = len(ofis)

    # dashboard
    prices = [r["price"] for r in L if r.get("price")]
    for col in range(1, 20): ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions["A"].width = 2.5
    ws.merge_cells("B1:S1"); t = ws["B1"]; t.value = "tap.az RADAR — Bazar Monitorinqi"
    t.font = Font(name=A, bold=True, size=20, color="FFFFFF"); t.alignment = Alignment(vertical="center", indent=1)
    for c in range(2, 20): ws.cell(row=1, column=c).fill = HF
    ws.row_dimensions[1].height = 40
    ws.merge_cells("B2:S2"); s = ws["B2"]
    s.value = f"Son skan: {run_ts} · {len(L)} aktiv elan · {len(news)} YENİ bu skanda"
    s.font = Font(name=A, size=10, italic=True, color="44546A"); s.alignment = Alignment(indent=1)
    for c in range(2, 20): ws.cell(row=2, column=c).fill = PatternFill("solid", fgColor="DDEBF7")

    def card(col, label, val, accent, big=20):
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 2)
        ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col + 2)
        lc = ws.cell(row=4, column=col, value=label); lc.font = Font(name=A, size=9, bold=True, color="FFFFFF")
        lc.alignment = Alignment(vertical="center", horizontal="center")
        vc = ws.cell(row=5, column=col, value=val); vc.font = Font(name=A, bold=True, size=big, color=accent)
        vc.alignment = Alignment(vertical="center", horizontal="center"); vc.number_format = "#,##0"
        th = Side(style="thin", color=accent)
        for c in range(col, col + 3):
            ws.cell(row=4, column=c).fill = PatternFill("solid", fgColor=accent)
            for rw in (4, 5, 6): ws.cell(row=rw, column=c).border = Border(left=th, right=th, top=th, bottom=th)
    card(2, "CƏMİ ELAN", len(L), "1F4E79")
    card(5, "🆕 YENİ", len(news), "2E7D32")
    card(8, "ORTA QİYMƏT", round(sum(prices) / len(prices)) if prices else 0, "C55A11")
    card(11, "🎮 GAMING", len(gaming), "6A1B9A")
    card(14, "💼 OFİS", len(ofis), "1565C0")

    def bar(title, col, ncount, anchor, color, pie=False):
        ch = PieChart() if pie else BarChart()
        if not pie: ch.type = "col"; ch.legend = None
        ch.title = title; ch.height = 6.6; ch.width = 12
        data = Reference(hs, min_col=col, min_row=1, max_row=1 + ncount)
        cats = Reference(hs, min_col=col - 1, min_row=2, max_row=1 + ncount)
        ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
        if not pie:
            try: ch.series[0].graphicalProperties.solidFill = color
            except Exception: pass
        ws.add_chart(ch, anchor)
    bar("Qiymət aralığı üzrə", 2, len(band_order), "B8", "1F4E79")
    bar("Brend üzrə", 5, min(10, len(brand_c)), "J8", "548235")
    bar("Gaming vs Ofis", 8, 2, "B24", None, pie=True)

    ws.cell(row=40, column=2, value="Dəyər balı = Spec ÷ Qiymət × 1000 · 🟩 yaşıl = ən yaxşı dəyər · 🆕/yaşıl sətir = bu skanda yeni · narıncı sətir = şəxsi satıcı (zəmanətsiz)").font = Font(name=A, italic=True, size=9, color="808080")
    ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1

    # move 🆕 Yeni right after Xülasə (it already is index1). Save.
    wb.save(path)
    return path
